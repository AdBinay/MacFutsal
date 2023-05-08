from typing import Any, Dict
from django.forms.models import BaseModelForm
from django.http import HttpResponse
from django.shortcuts import render
from django.views import generic 
from . import models,mixins,forms
from django.urls import reverse_lazy,reverse
from django.contrib.auth.mixins import LoginRequiredMixin,UserPassesTestMixin
from django.shortcuts import get_object_or_404

class HomeView(generic.ListView):
    template_name = "pages/home.html"
    model = models.Post

class CreatePostView(mixins.IsAdminOnlyMixin,generic.CreateView):
    template_name = "futsal/forms/post-create.html"
    model = models.Post
    fields = ["title","post_image","description"]
    success_url = reverse_lazy("home")
    
    def form_valid(self, form):
        post = form.save(commit=False)
        post.owner = self.request.user
        return super().form_valid(form)

class RequestCorApplicaiton(LoginRequiredMixin,generic.View):

    def get_object(self):
        return get_object_or_404(models.Post,id=self.kwargs.get("pk"))

    def get(self,request,*args, **kwargs):
        post = self.get_object()
        queryset = models.Coordinator.objects.filter(post=post)
        
        return render(request,"partials/cor-list.html",{"object_list":queryset})
    
    def post(self,request,*args, **kwargs):
        post = self.get_object()
        if self.request.user not in post.cordinator_in_review():
            post.add_cor_application(self.request.user)
            return HttpResponse("Redo Cordinator")
        else:
            post.remove_cor_application(self.request.user)
            return HttpResponse("Request Cordinator")
    
class RequestGMApplicaiton(LoginRequiredMixin,generic.View):
    
    def get_object(self):
        return get_object_or_404(models.Post,pk=self.kwargs.get("pk"))

    def get(self,request,*args, **kwargs):
        post = self.get_object()
        queryset = models.GameManager.objects.filter(post=post)
        return render(request,"partials/gm-list.html",{"object_list":queryset})
    
    def post(self,request,*args, **kwargs):
        post = self.get_object()
        if self.request.user not in post.gm_in_review():
            post.add_gm_application(self.request.user)
            return HttpResponse("Redo GM")
        else:
            post.remove_gm_application(self.request.user)
            return HttpResponse("Request GM")
    
class SelectCorView(LoginRequiredMixin,generic.View):
    template_name = "futsal/partials/select-cor-response.html"
    def get_post(self):
        return get_object_or_404(models.Post,pk=self.kwargs.get("pk"))
    
    def get_cor_instance(self):
        post = self.get_post()
        return get_object_or_404(models.Coordinator,post=post,pk=self.kwargs.get("pk_cor"))

    def post(self,request,*args, **kwargs):
        post = self.get_post()
        instance = self.get_cor_instance()

        if post.is_cordinator(instance.coordinator):
            post.unset_cordinator()
        else:
            post.set_cordinator(instance.coordinator)
            
        queryset = models.Coordinator.objects.filter(post=post)
        return render(request,"partials/cor-list.html",{"object_list":queryset}) 
    
    
class SelectGMView(LoginRequiredMixin,generic.View):
    template_name = "futsal/partials/select-gm-response.html"
    def get_post(self):
        return get_object_or_404(models.Post,id=self.kwargs.get("pk"))
    
    def gm_instance(self):
        return get_object_or_404(models.GameManager,id=self.kwargs.get("pk_gm"))

    def post(self,request,*args, **kwargs):
        post = self.get_post()
        instance = self.gm_instance()
        if post.is_gm(instance.game_manager):
            post.unset_gm()
        else:
            post.set_gm(instance.game_manager)
        print(post.get_gm_user())        
        queryset = models.GameManager.objects.filter(post=post)
        return render(request,"partials/gm-list.html",{"object_list":queryset})

class TeamCreateView(mixins.IsCordinatorMixin,generic.CreateView):
    template_name = "futsal/forms/team-create.html"
    model = models.Team
    form_class = forms.TeamCreateForm
    context_object_name = "team_form"

    def get_context_data(self, **kwargs):
        kwargs = super().get_context_data(**kwargs)
        kwargs[self.context_object_name] = kwargs["form"]
        return kwargs
    
    def form_valid(self, form):
        post = self.get_post()
        team = form.save(commit=False)
        team.post = post
        team.save()
        return super().form_valid(form)

    def get_success_url(self):
        post = self.get_post()
        return reverse('team-manager',kwargs={"pk":post.id})


class PlayerCreateView(mixins.IsCordinatorMixin,generic.CreateView):
    template_name = "futsal/forms/player-create.html"
    model = models.Players
    form_class = forms.PlayerForm
    context_object_name = "player_form"

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs.update({"post":self.get_post()})
        print(kwargs)
        return kwargs
    
    def get_context_data(self, **kwargs):
        kwargs = super().get_context_data(**kwargs)
        kwargs[self.context_object_name] = kwargs["form"]
        return kwargs
    
    def get_success_url(self):
        post = self.get_post()
        return reverse('team-manager',kwargs={"pk":post.id}) 

    def form_valid(self, form):
        post = self.get_post()
        player = form.save(commit=False)
        player.post = post
        player.save()
        return super().form_valid(form)


class TeamManagerView(mixins.IsCordinatorMixin,generic.ListView):
    template_name = "pages/team-manager.html"
    model = models.Players

    def get_queryset(self):
        post = self.get_post()
        return super().get_queryset().filter(post=post)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["team_form"] =  forms.TeamCreateForm()
        context["player_form"] =  forms.PlayerForm(**{"post":self.get_post()})
        return context

class CorUpdatePlayer(mixins.IsCordinatorMixin,generic.UpdateView):
    template_name = "forms/update-player-form.html"
    pk_url_kwarg = "pk_player"
    model = models.Players
    form_class = forms.PlayerForm

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs.update({"post":self.get_post()})
        print(kwargs)
        return kwargs
    
    def get_context_data(self, **kwargs):
        instance = self.get_object()
        context = super().get_context_data(**kwargs)
        context["player"] = instance
        return context
    
    def get_success_url(self):
        post = self.get_post()
        return reverse('team-manager',kwargs={"pk":post.id}) 

class CorDeletePlayer(mixins.IsCordinatorMixin,generic.DeleteView):
    http_method_names = "post"
    pk_url_kwarg = "pk_player"
    model = models.Players
    
    def get_success_url(self):
        post = self.get_post()
        team = self.get_object()
        return reverse('team-manager',kwargs={"pk":post.id})

class MatchCreateView(mixins.IsGameManagerMixin,generic.CreateView):
    form_class = forms.MatchForm
    template_name = "futsal/forms/match-create.html"
    model = models.Match
    context_object_name = "match_form"
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["post"] = self.get_post()
        return kwargs

    def get_context_data(self, **kwargs):
        kwargs = super().get_context_data(**kwargs)
        kwargs[self.context_object_name] = kwargs["form"]
        return kwargs
    
    def get_success_url(self):
        post = self.get_post()
        return reverse('match-manager',kwargs={"pk":post.id}) 

    def form_valid(self, form):
        post = self.get_post()
        match = form.save(commit=False)
        match.post = post
        match.save()
        return super().form_valid(form)

class MatchManagerView(mixins.IsGameManagerMixin,generic.ListView):
    template_name = "pages/match-manager.html"
    model = models.Match
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["match_form"] = forms.MatchForm(**{"post":self.get_post()}) 
        return context

    def get_queryset(self):
        post = self.get_post()
        return super().get_queryset().filter(post=post)
    
class MatchUpdateView(mixins.IsGameManagerMixin,generic.UpdateView):
    template_name = "forms/update-match-form.html"
    pk_url_kwarg = "pk_match"
    model = models.Match
    form_class = forms.MatchFormUpdate

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["post"] = self.get_post()
        return kwargs

    def get_context_data(self, **kwargs):
        instance = self.get_object()
        context = super().get_context_data(**kwargs)
        context["match"] = instance
        return context
    
    def get_success_url(self):
        post = self.get_post()
        return reverse('match-manager',kwargs={"pk":post.id}) 

class TeamDetailView(LoginRequiredMixin,mixins.GetPostMixin,generic.DetailView):
    template_name = "pages/team-detail.html"
    model = models.Team
    context_object_name = "team"
    pk_url_kwarg = "pk_team"

class TeamUpdateView(mixins.IsCordinatorMixin,generic.UpdateView):
    form_class = forms.TeamUpdateForm
    template_name = "futsal/forms/team-update.html"
    pk_url_kwarg = "pk_team"
    context_object_name = "team"
    model = models.Team

    def get_queryset(self):
        post = self.get_post()
        return super().get_queryset().filter(post=post)

    def get_success_url(self):
        post = self.get_post()
        team = self.get_object()
        return reverse('team-detail',kwargs={"pk":post.id,"pk_team":team.id}) 

class TeamDeleteView(mixins.IsCordinatorMixin,generic.DeleteView):
    http_method_names=["post"]
    pk_url_kwarg = "pk_team"
    context_object_name = "team"
    model = models.Team

    def get_success_url(self):
        post = self.get_post()
        team = self.get_object()
        return reverse('team-manager',kwargs={"pk":post.id}) 

class MatchDetailView(mixins.GetPostMixin,generic.ListView):
    template_name = "pages/match-detail.html"
    model = models.Match
    pk_url_kwarg = "pk_match"
    context_object_name = "match"

    def get_match(self):
        if not hasattr(self,"match_instance"):
            self.match_instance = get_object_or_404(self.model,id=self.kwargs.get(str(self.pk_url_kwarg)))
        return self.match_instance

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context[self.context_object_name] = self.get_match()
        return context


class GMUpdatePlayer(mixins.IsGameManagerMixin,generic.UpdateView):

    template_name = "forms/gm-update-player-form.html"
    pk_url_kwarg = "pk_player"
    model = models.Players
    form_class = forms.MatchPlayerForm
    
    def get_context_data(self, **kwargs):
        instance = self.get_object()
        context = super().get_context_data(**kwargs)
        context["player"] = instance
        return context
    
    def form_valid(self, form):
        form.save(commit=True)
        instance = self.get_object()
        queryset = instance.team.all_players()
        total_score = queryset.aggregate(models.models.Sum('score'))["score__sum"]
        instance.team.score = total_score
        instance.team.save()

        print(total_score)
        return super().form_valid(form)
    
    def get_success_url(self):
        # post = self.get_post()
        # instance = self.get_object()
        # return reverse('match-detail',kwargs={"pk":post.id,"pk_match":instance.id}) 
        return self.request.META.get('HTTP_REFERER')
    
class ViewTeamList(mixins.GetPostMixin,generic.ListView):
    template_name = "futsal/pages/team-list.html"
    model = models.Team

class ViewMatchList(mixins.GetPostMixin,generic.ListView):
    template_name = "futsal/pages/match-list.html"
    model = models.Match



###tie-sheet generations

from django.shortcuts import render
from django.http import HttpResponse
from django.core.files.base import ContentFile
from django.core.files import File
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, borders
from .models import Team
import pandas as pd

# def tie_sheet(request):
#     # Load the Excel template
#     workbook = load_workbook(filename='static/excel/Tie-sheet.xlsx')
    
#     # Get the active worksheet
#     worksheet = workbook.active
    
#     # # Create merged cell for POLL1 and POLL2 columns and set heading
#     # worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=3)
#     # worksheet.cell(row=1, column=2, value='POLL1')
#     # worksheet.cell(row=1, column=4, value='POLL2')
    
#     # Get the teams from the database
#     teams = Team.objects.all().order_by('?')
    
#     # Fill teams in D5, D10, D15, and D20 columns
#     count = 1
#     for team in teams:
#         if count == 1:
#             cell = worksheet.cell(row=5, column=4, value=team.title)
#         elif count == 2:
#             cell = worksheet.cell(row=10, column=4, value=team.title)
#         elif count == 3:
#             cell = worksheet.cell(row=15, column=4, value=team.title)
#         elif count == 4:
#             cell = worksheet.cell(row=20, column=4, value=team.title)
#         cell.border = borders.Border(left=borders.Side(style='thin'), right=borders.Side(style='thin'), top=borders.Side(style='thin'), bottom=borders.Side(style='thin'))
#         count += 1
    
#     # Add spacing between rows and columns
#     for row in worksheet.iter_rows(min_row=1, max_row=1):
#         for cell in row:
#             cell.font = Font(size=12, bold=True)
#             cell.alignment = Alignment(horizontal='center')
#             #cell.fill = PatternFill(start

    
#     for col in worksheet.iter_cols(min_col=2, max_col=4):
#         for cell in col:
#             cell.font = Font(size=12, bold=True)
#             cell.alignment = Alignment(horizontal='center')
    
#     # Create a response with the filled Excel file
#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = 'attachment; filename=teams.xlsx'
#     workbook.save(response)
    
#     return response

def tie_sheet(request):
    # Get the number of teams
    num_teams = Team.objects.count()
    
    # Select the appropriate template file based on the number of teams
    if num_teams == 4:
        template_file = 'static/excel/Tie-sheet.xlsx'
    elif num_teams == 5:
        template_file = 'static/excel/Tie-sheet5.xlsx'
    elif num_teams == 6:
        template_file = 'static/excel/Tie-sheet6.xlsx'
    elif num_teams == 8:
        template_file = 'static/excel/Tie-sheet8.xlsx'
    else:
        # Handle invalid number of teams
        return HttpResponse('Invalid number of teams')
    
    # Load the Excel template
    workbook = load_workbook(filename=template_file)
    
    # Get the active worksheet
    worksheet = workbook.active
    
    # Fill teams in the appropriate columns based on the number of teams
    if num_teams == 4:
        column = 4
        rows = [5, 10, 15, 20]
    elif num_teams == 5:
        column = 4
        rows = [4, 8, 12, 16, 20]
    elif num_teams == 6:
        column = 4
        rows = [3, 7, 11, 15, 19, 23]
    elif num_teams == 8:
        column = 4
        rows = [3, 6, 9, 12, 16, 19, 22, 25]
    else:
        # Handle invalid number of teams
        return HttpResponse('Invalid number of teams')
    
    # Fill teams in the appropriate cells
    count = 0
    for team in Team.objects.order_by('?'):
        cell = worksheet.cell(row=rows[count], column=column, value=team.title)
        cell.border = borders.Border(left=borders.Side(style='thin'), right=borders.Side(style='thin'), top=borders.Side(style='thin'), bottom=borders.Side(style='thin'))
        count += 1
    
    # Add spacing between rows and columns
    for row in worksheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = Font(size=12, bold=True)
            cell.alignment = Alignment(horizontal='center')
    
    for col in worksheet.iter_cols(min_col=2, max_col=column):
        for cell in col:
            cell.font = Font(size=12, bold=True)
            cell.alignment = Alignment(horizontal='center')
    
    # Create a response with the filled Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=teams.xlsx'
    workbook.save(response)
    
    return response
def tie_sheet(request):
    # Get the number of teams
    num_teams = Team.objects.count()
    
    # Select the appropriate template file based on the number of teams
    if num_teams == 4:
        template_file = 'static/excel/Tie-sheet.xlsx'
    elif num_teams == 5:
        template_file = 'static/excel/Tie-sheet5.xlsx'
    elif num_teams == 6:
        template_file = 'static/excel/Tie-sheet6.xlsx'
    elif num_teams == 8:
        template_file = 'static/excel/Tie-sheet8.xlsx'
    else:
        # Handle invalid number of teams
        return HttpResponse('Invalid number of teams')
    
    # Load the Excel template
    workbook = load_workbook(filename=template_file)
    
    # Get the active worksheet
    worksheet = workbook.active
    
    # Fill teams in the appropriate columns based on the number of teams
    if num_teams == 4:
        column = 4
        rows = [5, 10, 15, 20]
    elif num_teams == 5:
        column = 4
        rows = [5,10,15,20,25]
    elif num_teams == 6:
        column = 4
        rows = [5,10,15,20,25,30]
    elif num_teams == 8:
        column = 4
        rows = [5,10,15,20,25,30,35]
    else:
        # Handle invalid number of teams
        return HttpResponse('Invalid number of teams')
    
    # Fill teams in the appropriate cells
    count = 0
    for team in Team.objects.order_by('?'):
        cell = worksheet.cell(row=rows[count], column=column, value=team.title)
        cell.border = borders.Border(left=borders.Side(style='thin'), right=borders.Side(style='thin'), top=borders.Side(style='thin'), bottom=borders.Side(style='thin'))
        count += 1
    
    # Add spacing between rows and columns
    for row in worksheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = Font(size=12, bold=True)
            cell.alignment = Alignment(horizontal='center')
    
    for col in worksheet.iter_cols(min_col=2, max_col=column):
        for cell in col:
            cell.font = Font(size=12, bold=True)
            cell.alignment = Alignment(horizontal='center')
    
    # Create a response with the filled Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=teams.xlsx'
    workbook.save(response)
    
    return response
